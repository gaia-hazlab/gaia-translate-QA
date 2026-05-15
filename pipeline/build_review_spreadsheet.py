#!/usr/bin/env python3
"""
build_review_spreadsheet.py — produce the reviewer xlsx from eval_dataset_v3.json.

The output xlsx has four tabs:

1. **Instructions** — meta + reviewer-identification fields.
2. **QAs** — one row per QA. Includes:
   - Identifier columns (id, disciplines, query_type, difficulty, status)
   - Input (prompt, document type/source/title)
   - Expected output as separate columns: concept_matches, method_matches,
     phenomenon_matches, translation_matches, vocab collisions, refusals,
     themes, must_not_say
   - Reviewer input columns (empty): per-section scores S1-S8 against the
     rubric in eval_platform/, overall comment, confidence, time, flag, status
3. **Rubric** — the 8-criterion scoring rubric for reference.
4. **Progress** — auto-summary formulas.

Usage
-----
    python pipeline/build_review_spreadsheet.py
    python pipeline/build_review_spreadsheet.py --in eval_dataset/eval_dataset_v3.json --out eval_dataset/gaia_translator_eval_review_v3.xlsx

Author: Denolle Group.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ============================================================================
# STYLES (match eval_platform/templates/scoring_template.xlsx)
# ============================================================================

ARIAL = "Arial"
ACCENT = "1F4E79"
HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
INPUT_BLUE = "0000FF"
FORMULA_BLACK = "000000"
WARN_YELLOW = "FFF2CC"
GOOD_GREEN = "C6EFCE"
ROW_ALT = "F2F2F2"

thin = Side(border_style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_cell(cell, text, fill=HEADER_BG, fg=HEADER_FG):
    cell.value = text
    cell.font = Font(name=ARIAL, size=11, bold=True, color=fg)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_all


def text_cell(cell, value, wrap=True, bold=False, color_hex=FORMULA_BLACK, fill=None):
    cell.value = value
    cell.font = Font(name=ARIAL, size=10, bold=bold, color=color_hex)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    cell.border = border_all
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def input_cell(cell, value=None):
    if value is not None:
        cell.value = value
    cell.font = Font(name=ARIAL, size=10, color=INPUT_BLUE)
    cell.fill = PatternFill("solid", fgColor="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    cell.border = border_all


def formula_cell(cell, value):
    cell.value = value
    cell.font = Font(name=ARIAL, size=10, color=FORMULA_BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.border = border_all


# ============================================================================
# TAB BUILDERS
# ============================================================================

def build_instructions_tab(ws, total_qas: int, stats: Dict) -> None:
    ws["A1"] = "Gaia translator — v3 eval review"
    ws["A1"].font = Font(name=ARIAL, size=18, bold=True, color=ACCENT)

    ws["A3"] = "Reviewer name"
    ws["A3"].font = Font(name=ARIAL, size=11, bold=True)
    input_cell(ws["B3"], "(your name)")
    ws["A4"] = "Primary discipline"
    ws["A4"].font = Font(name=ARIAL, size=11, bold=True)
    input_cell(ws["B4"], "(your primary)")
    ws["A5"] = "Total QAs assigned"
    text_cell(ws["B5"], total_qas)
    ws["A6"] = "Calibration round deadline"
    input_cell(ws["B6"], "(yyyy-mm-dd)")
    ws["A7"] = "Production scoring deadline"
    input_cell(ws["B7"], "(yyyy-mm-dd)")

    notes = [
        "",
        "How this sheet works",
        "",
        "Each row in the 'QAs' tab is one QA. Columns A-U describe the QA (A-H are the v3.1 classification: ID, status, tier, disciplines, query type, translation task types, compound coupling, difficulty; I-L are the prompt + document; M-U are the expected output incl. failure modes tested). Columns V-AC are the 'S1...S8' score inputs (1-5, leave blank if you can't judge), followed by mean / comment / confidence / status (AD-AJ).",
        "Blue text = your input. Type scores and comments here.",
        "Black text = formula (don't edit).",
        "Yellow cells = mandatory comment if score is ≤2 or ≥5.",
        "",
        "Eval set stats:",
        f"  Total QAs: {total_qas}",
        f"  By query type: {stats.get('by_query_type', {})}",
        f"  By difficulty: {stats.get('by_difficulty', {})}",
        f"  By discipline: {stats.get('by_discipline', {})}",
        f"  By disc-span:  {stats.get('by_span', {})}",
        "",
        "See eval_platform/templates/reviewer_instructions.docx for the full rubric and review process.",
    ]
    row = 9
    for line in notes:
        c = ws.cell(row=row, column=1, value=line)
        c.font = Font(name=ARIAL, size=11, bold=(line == "How this sheet works"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40


def build_qas_tab(ws, qas: List[Dict]) -> None:
    """The main scoring tab — one row per QA."""
    # Column layout (v3.1 — adds tier, translation_task_types, compound_coupling, failure_modes_tested)
    columns = [
        ("QA_ID", 14),
        ("Status", 12),
        ("Tier", 9),                    # v3.1
        ("Disciplines", 22),
        ("Query type", 18),
        ("Translation task types", 28), # v3.1
        ("Compound coupling", 28),      # v3.1
        ("Difficulty", 10),
        ("Prompt", 50),
        ("Doc type", 10),
        ("Doc source", 28),
        ("Doc title", 24),
        # Expected output
        ("Expected concept_matches", 28),
        ("Expected method_matches", 28),
        ("Expected phenomenon_matches", 28),
        ("Expected translation_matches", 28),
        ("Expected vocab collisions", 22),
        ("Expected refusals", 24),
        ("Failure modes tested", 28),   # v3.1
        ("Expected response themes", 60),
        ("Must not say", 40),
        # Score inputs (blue)
        ("S1 Tech accuracy", 10),
        ("S2 Citation discipline", 11),
        ("S3 Vocab precision", 10),
        ("S4 Cross-disc integration", 11),
        ("S5 Refusal correctness", 11),
        ("S6 Completeness", 10),
        ("S7 Presentation", 10),
        ("S8 Overall usefulness", 11),
        # Derived
        ("Mean score (auto)", 12),
        ("Overall comment (REQUIRED)", 50),
        ("Confidence (1-5)", 11),
        ("Skill files to revise", 28),
        ("Flag for discussion (Y/N)", 12),
        ("Time to score (min)", 11),
        ("Status (auto)", 14),
    ]

    for idx, (h, w) in enumerate(columns, start=1):
        header_cell(ws.cell(row=1, column=idx), h)
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.row_dimensions[1].height = 46
    ws.freeze_panes = "I2"  # freeze through column H (so prompt scrolls with scores)

    # New column index layout (v3.1; +4 columns vs. v3.0)
    SCORE_COLS = list(range(22, 30))  # S1..S8 — shifted by +4
    MEAN_COL = 30
    OVERALL_COMMENT = 31
    CONFIDENCE = 32
    SKILL_FILES = 33
    FLAG = 34
    TIME_MIN = 35
    STATUS = 36

    # Data rows
    for i, qa in enumerate(qas):
        r = i + 2
        ws.row_dimensions[r].height = 220

        text_cell(ws.cell(row=r, column=1), qa["id"], bold=True)
        text_cell(ws.cell(row=r, column=2), qa["status"])
        text_cell(ws.cell(row=r, column=3), qa.get("tier", ""), bold=True)
        text_cell(ws.cell(row=r, column=4), ", ".join(qa["primary_disciplines"]))
        text_cell(ws.cell(row=r, column=5), qa["query_type"])
        text_cell(ws.cell(row=r, column=6), "\n".join(qa.get("translation_task_types", [])))
        text_cell(ws.cell(row=r, column=7), "\n".join(qa.get("compound_coupling", [])))
        text_cell(ws.cell(row=r, column=8), qa["difficulty"])
        text_cell(ws.cell(row=r, column=9), qa["prompt"])
        doc = qa.get("input_document", {})
        text_cell(ws.cell(row=r, column=10), doc.get("type", "none"))
        text_cell(ws.cell(row=r, column=11), doc.get("source") or "")
        text_cell(ws.cell(row=r, column=12), doc.get("title") or "")

        exp = qa.get("expected_output", {})
        text_cell(ws.cell(row=r, column=13), "\n".join(exp.get("concept_matches", [])))
        text_cell(ws.cell(row=r, column=14), "\n".join(exp.get("method_matches", [])))
        text_cell(ws.cell(row=r, column=15), "\n".join(exp.get("phenomenon_matches", [])))
        text_cell(ws.cell(row=r, column=16), "\n".join(exp.get("translation_matches", [])))
        text_cell(ws.cell(row=r, column=17), ", ".join(exp.get("vocabulary_collisions_flagged", [])))
        text_cell(ws.cell(row=r, column=18), ", ".join(exp.get("refusals_or_caveats_expected", [])))
        text_cell(ws.cell(row=r, column=19), "\n".join(exp.get("failure_modes_tested", [])))
        themes = exp.get("user_specific_response_themes", [])
        text_cell(ws.cell(row=r, column=20),
                  "\n".join(f"• {t}" for t in themes))
        must_not = exp.get("must_not_say", [])
        text_cell(ws.cell(row=r, column=21),
                  "\n".join(f"✗ {t}" for t in must_not))

        # Score input cells (blue)
        for sc_idx in SCORE_COLS:
            input_cell(ws.cell(row=r, column=sc_idx))

        # Mean formula
        cells = [f"{get_column_letter(c)}{r}" for c in SCORE_COLS]
        sum_expr = "+".join([f"IFERROR(VALUE({c}),0)" for c in cells])
        count_expr = "+".join([f"IF(ISNUMBER({c}),1,0)" for c in cells])
        formula_cell(ws.cell(row=r, column=MEAN_COL),
                     f"=IFERROR(({sum_expr})/({count_expr}),\"\")")

        # Reviewer inputs
        input_cell(ws.cell(row=r, column=OVERALL_COMMENT))
        input_cell(ws.cell(row=r, column=CONFIDENCE))
        input_cell(ws.cell(row=r, column=SKILL_FILES))
        input_cell(ws.cell(row=r, column=FLAG))
        input_cell(ws.cell(row=r, column=TIME_MIN))

        # Status formula
        score_count = "+".join([f"IF(ISNUMBER({c})+ISTEXT({c}),1,0)" for c in cells])
        overall_filled = f"IF({get_column_letter(OVERALL_COMMENT)}{r}<>\"\",1,0)"
        status_formula = (
            f"=IF(AND(({score_count})>=4,{overall_filled}=1),\"COMPLETE\","
            f"IF(({score_count})=0,\"NOT STARTED\",\"IN PROGRESS\"))"
        )
        formula_cell(ws.cell(row=r, column=STATUS), status_formula)

    # Data validation: scores 1-5 or "N/A"
    dv_score = DataValidation(type="list", formula1='"1,2,3,4,5,N/A"',
                              allow_blank=True, errorTitle="Invalid score")
    for sc_idx in SCORE_COLS:
        letter = get_column_letter(sc_idx)
        dv_score.add(f"{letter}2:{letter}{len(qas)+1}")
    ws.add_data_validation(dv_score)

    # Confidence
    dv_conf = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_conf.add(f"{get_column_letter(CONFIDENCE)}2:{get_column_letter(CONFIDENCE)}{len(qas)+1}")
    ws.add_data_validation(dv_conf)

    # Flag
    dv_flag = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_flag.add(f"{get_column_letter(FLAG)}2:{get_column_letter(FLAG)}{len(qas)+1}")
    ws.add_data_validation(dv_flag)

    # Conditional formatting: warn yellow when score is 1, 2, 5 and overall comment is empty
    for sc_idx in SCORE_COLS:
        sc_letter = get_column_letter(sc_idx)
        rule = FormulaRule(
            formula=[f"AND(OR({sc_letter}2=1,{sc_letter}2=2,{sc_letter}2=5),"
                     f"{get_column_letter(OVERALL_COMMENT)}2=\"\")"],
            fill=PatternFill("solid", fgColor=WARN_YELLOW),
        )
        ws.conditional_formatting.add(f"{sc_letter}2:{sc_letter}{len(qas)+1}", rule)

    # Status: green when complete
    ws.conditional_formatting.add(
        f"{get_column_letter(STATUS)}2:{get_column_letter(STATUS)}{len(qas)+1}",
        CellIsRule(operator="equal", formula=['"COMPLETE"'],
                   fill=PatternFill("solid", fgColor=GOOD_GREEN)),
    )


def build_rubric_tab(ws) -> None:
    ws["A1"] = "Scoring rubric — 8 criteria, 1-5 Likert"
    ws["A1"].font = Font(name=ARIAL, size=16, bold=True, color=ACCENT)
    ws.merge_cells("A1:E1")

    headers = ["Criterion", "Definition", "1 - poor", "3 - acceptable", "5 - excellent"]
    widths = [28, 50, 30, 30, 30]
    for idx, (h, w) in enumerate(zip(headers, widths), start=1):
        header_cell(ws.cell(row=3, column=idx), h)
        ws.column_dimensions[get_column_letter(idx)].width = w

    rubric = [
        ("S1. Technical accuracy",
         "Are the factual claims in the answer correct?",
         "Wrong claim that would mislead.",
         "Mostly correct; minor inaccuracy.",
         "All claims correct; uncertainty quantified."),
        ("S2. Citation discipline",
         "Are cited works real, DOIs accurate, claims supported?",
         "Fabricated citation; hard fail.",
         "Citations real; coverage uneven.",
         "Every claim cited; coverage matches specificity."),
        ("S3. Vocabulary precision",
         "Discipline-specific language correct; vocabulary collisions handled?",
         "Vocabulary collision treated as equivalence.",
         "Language correct; collisions disambiguated when prompted.",
         "Language fully correct; collisions disambiguated up front."),
        ("S4. Cross-discipline integration (N/A single-disc)",
         "Does the answer correctly bridge disciplines?",
         "Forces analogy that fails scrutiny.",
         "Identifies bridge but doesn't exploit it.",
         "Right bridge; both sides worked through; flags failure modes."),
        ("S5. Refusal correctness (N/A no refusal opportunity)",
         "Does the chatbot refuse forced analogies/fabrications?",
         "Failed to refuse; fabricated citation.",
         "Hedged but didn't name the pattern.",
         "Cleanly refused with named pattern + alternative."),
        ("S6. Completeness",
         "Address what was asked? Important parts missing?",
         "Major aspects unaddressed.",
         "Main question addressed; peripheral missing.",
         "All aspects addressed."),
        ("S7. Presentation",
         "Clarity, structure, technical level.",
         "Disorganized; over- or under-technical.",
         "Reasonably structured.",
         "Well-structured; technical level correct."),
        ("S8. Overall usefulness",
         "Holistic judgment: would a researcher find this helpful?",
         "Not useful; would mislead.",
         "Useful as starting point; need to verify.",
         "Directly useful; would quote in own work."),
    ]
    for i, row_data in enumerate(rubric):
        r = 4 + i
        fill = ROW_ALT if i % 2 == 1 else "FFFFFF"
        for col_idx, val in enumerate(row_data, start=1):
            text_cell(ws.cell(row=r, column=col_idx), val, wrap=True, bold=(col_idx == 1),
                      fill=fill)
        ws.row_dimensions[r].height = 80


def build_progress_tab(ws, total_qas: int) -> None:
    ws["A1"] = "Your scoring progress"
    ws["A1"].font = Font(name=ARIAL, size=16, bold=True, color=ACCENT)
    ws.merge_cells("A1:D1")

    # v3.1: column positions shifted by +4 due to new dimension columns.
    # Status col 36 → AJ; Mean col 30 → AD; Time col 35 → AI; Flag col 34 → AH.
    items = [
        ("Total QAs", f"=COUNTA(QAs!A2:A{total_qas+1})"),
        ("Complete", f"=COUNTIF(QAs!AJ2:AJ{total_qas+1},\"COMPLETE\")"),
        ("In progress", f"=COUNTIF(QAs!AJ2:AJ{total_qas+1},\"IN PROGRESS\")"),
        ("Not started", f"=COUNTIF(QAs!AJ2:AJ{total_qas+1},\"NOT STARTED\")"),
        ("Percent complete", f"=IFERROR(B3/B2,0)"),
        ("Mean score across complete", f"=IFERROR(AVERAGEIF(QAs!AJ2:AJ{total_qas+1},\"COMPLETE\",QAs!AD2:AD{total_qas+1}),\"\")"),
        ("Mean time per QA (min)", f"=IFERROR(AVERAGEIF(QAs!AJ2:AJ{total_qas+1},\"COMPLETE\",QAs!AI2:AI{total_qas+1}),\"\")"),
        ("Flagged for discussion", f"=COUNTIF(QAs!AH2:AH{total_qas+1},\"Y\")"),
    ]
    for i, (label, formula) in enumerate(items):
        r = 2 + i
        ws.cell(row=r, column=1, value=label).font = Font(name=ARIAL, size=11, bold=True)
        ws.cell(row=r, column=2, value=formula).font = Font(name=ARIAL, size=11)
    ws.cell(row=6, column=2).number_format = "0.0%"
    ws.cell(row=7, column=2).number_format = "0.00"
    ws.cell(row=8, column=2).number_format = "0.0"

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22


# ============================================================================
# MAIN
# ============================================================================

def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Build reviewer xlsx from eval_dataset_v3.json.")
    parser.add_argument(
        "--in", dest="input_path",
        type=Path,
        default=Path("eval_dataset/eval_dataset_v3.json"),
        help="Input JSON eval set.",
    )
    parser.add_argument(
        "--out", dest="output_path",
        type=Path,
        default=Path("eval_dataset/gaia_translator_eval_review_v3.xlsx"),
        help="Output xlsx.",
    )
    parser.add_argument(
        "--corpus-root",
        type=Path,
        default=Path(__file__).resolve().parent.parent,
        help="Root of gaia-translate-QA repo.",
    )
    args = parser.parse_args(argv)

    in_path = args.corpus_root / args.input_path if not args.input_path.is_absolute() else args.input_path
    out_path = args.corpus_root / args.output_path if not args.output_path.is_absolute() else args.output_path

    if not in_path.exists():
        print(f"ERROR: input {in_path} does not exist. Run validate_eval_set.py first.",
              file=sys.stderr)
        return 2

    with in_path.open() as f:
        data = json.load(f)

    qas = data.get("qas", [])
    stats = data.get("stats", {})
    if not qas:
        print(f"ERROR: no QAs found in {in_path}.", file=sys.stderr)
        return 2

    wb = Workbook()
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    build_instructions_tab(ws_instructions, total_qas=len(qas), stats=stats)

    ws_qas = wb.create_sheet("QAs")
    build_qas_tab(ws_qas, qas)

    ws_rubric = wb.create_sheet("Rubric")
    build_rubric_tab(ws_rubric)

    ws_progress = wb.create_sheet("Progress")
    build_progress_tab(ws_progress, total_qas=len(qas))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}  ({len(qas)} QAs)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
